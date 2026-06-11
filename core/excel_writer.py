"""
범용 발주서 엑셀 생성기
======================
업체별 OUTPUT_CONFIG(딕셔너리)와 데이터(행 리스트)를 받아서,
해당 업체 양식에 맞는 엑셀 파일을 생성합니다.

OUTPUT_CONFIG 구조 예시는 vendors/hanpoom_comfortlab.py 를 참고하세요.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)


def _fill(color_hex):
    return PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")


def write_order_excel(output_path, data_rows, output_config, footer_info=None, meta=None,
                       master_products=None):
    """
    output_path: 저장할 xlsx 경로
    data_rows: [{"상품코드": ..., "품번": ..., ... } , ...]  (output_config["headers"]의 키를 사용)
    output_config: 업체별 출력 양식 정의 (dict)
    footer_info: PDF에서 추출한 key-value 정보 (예: {"납품 장소": "...", "입고 담당자": "..."})
    meta: PDF에서 추출한 발주번호/발주일자 등 메타 정보
    master_products: (선택) 참고용 마스터 상품리스트 시트에 넣을 데이터
    """
    footer_info = footer_info or {}
    meta = meta or {}

    headers = output_config["headers"]
    header_fills = output_config.get("header_fills", {})
    column_alignment = output_config.get("column_alignment", {})
    number_formats = output_config.get("number_format_columns", {})
    formula_columns = output_config.get("formula_columns", {})
    sum_columns = output_config.get("sum_columns", [])
    sheet_name = output_config.get("sheet_name", "발주서")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 상단 메타 정보 (발주번호/발주일자/납기일자 등)이 있으면 한 줄로 기록
    start_row_offset = 0
    if meta:
        meta_text = "  |  ".join(f"{k}: {v}" for k, v in meta.items())
        ws.cell(row=1, column=1, value=meta_text).font = Font(bold=True, size=11)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        start_row_offset = 2  # 메타 정보 줄 + 빈 줄

    header_row_idx = 1 + start_row_offset

    # 헤더 행
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if header in header_fills:
            cell.fill = _fill(header_fills[header])

    # 데이터 행
    start_row = header_row_idx + 1
    for r_offset, row_data in enumerate(data_rows):
        r_idx = start_row + r_offset
        for col_idx, header in enumerate(headers, 1):
            if col_idx in formula_columns:
                value = formula_columns[col_idx].format(row=r_idx)
            else:
                value = row_data.get(header, "")
            cell = ws.cell(row=r_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

            align = column_alignment.get(col_idx)
            if align:
                cell.alignment = Alignment(horizontal=align)

            if col_idx in number_formats:
                cell.number_format = number_formats[col_idx]

    # 합계 행
    total_row_idx = start_row + len(data_rows)
    if sum_columns:
        for col_idx in sum_columns:
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(
                row=total_row_idx, column=col_idx,
                value=f"=SUM({col_letter}{start_row}:{col_letter}{total_row_idx - 1})"
            )
            cell.fill = _fill("FFFF00")
            cell.font = Font(bold=True)
            if col_idx in number_formats:
                cell.number_format = number_formats[col_idx]
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=total_row_idx, column=col_idx).border = THIN_BORDER

    # 하단 정보 필드 (PDF에서 추출한 footer_info 그대로 출력)
    footer_fields = output_config.get("footer_fields", [])
    if footer_fields:
        info_start_row = total_row_idx + 3
        r = info_start_row
        for field in footer_fields:
            label = field["label"]
            pdf_key = field.get("pdf_key", label)
            value = footer_info.get(pdf_key, "")
            ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            ws.cell(row=r, column=2, value=value)
            for c in range(1, output_config.get("footer_width", 5) + 1):
                ws.cell(row=r, column=c).border = THIN_BORDER
            r += 1

    # 열 너비 자동 조정 (대략적인 휴리스틱)
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, len(header) + 4)

    # 참고용 마스터 상품리스트 시트
    if master_products:
        ws_list = wb.create_sheet(title=output_config.get("master_sheet_name", "제품리스트"))
        if master_products:
            keys = list(master_products[0].keys())
            ws_list.append(keys)
            for p in master_products:
                ws_list.append([p.get(k, "") for k in keys])

    wb.save(output_path)
    return output_path
